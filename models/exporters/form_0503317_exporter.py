"""Экспортёр и валидатор для формы 0503317"""
import shutil
import os
import openpyxl
from openpyxl.styles import PatternFill
from typing import Dict, List, Any, Optional
import pandas as pd
from logger import logger
from models.constants.form_0503317_constants import Form0503317Constants, StyleConstants
from models.utils.form_utils import column_to_index, extract_original_value_from_cell
from models.calculators.form_0503317_calculator import Form0503317Calculator
from utils.numeric_utils import is_value_different


class Form0503317Exporter:
    """Экспортёр и валидатор для формы 0503317"""
    
    def __init__(self, constants: Form0503317Constants, calculator: Form0503317Calculator):
        """
        Args:
            constants: Константы формы 0503317
            calculator: Калькулятор для пересчета сумм
        """
        self.constants = constants
        self.calculator = calculator
        self.show_error_values = True
    
    def export_validation(
        self, 
        original_file_path: str, 
        output_file_path: str, 
        form_data: Dict[str, Any],
        calculated_deficit_proficit: Optional[Dict[str, Dict[str, float]]] = None
    ) -> str:
        """Экспорт формы с проверкой
        
        Args:
            original_file_path: Путь к исходному файлу
            output_file_path: Путь для сохранения результата
            form_data: Словарь с данными формы (с расчетными значениями)
            calculated_deficit_proficit: Дефицит/профицит (опционально)
        
        Returns:
            Путь к сохраненному файлу
        
        Raises:
            FileNotFoundError: Если исходный файл не существует
        """
        # Проверяем существование исходного файла
        if not original_file_path or not os.path.exists(original_file_path):
            error_msg = f"Исходный файл не найден: {original_file_path}. Экспорт с валидацией требует исходный Excel файл."
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)
        
        # Нормализуем пути
        src = os.path.abspath(original_file_path)
        dst = os.path.abspath(output_file_path)

        # Логируем пути для диагностики
        logger.debug(f"Экспорт с проверкой: copy2 src='{src}', dst='{dst}'")

        # Если пользователь экспортирует в тот же самый файл, что и исходный,
        # копировать его "сам в себя" нельзя – просто работаем напрямую с ним.
        if os.path.normcase(src) == os.path.normcase(dst):
            wb = openpyxl.load_workbook(src)
        else:
            # Копируем исходный файл в выходной
            try:
                shutil.copy2(src, dst)
            except PermissionError as e:
                # Логируем детальную информацию, чтобы понять, какой файл заблокирован
                logger.error(
                    "PermissionError при копировании файла.\n"
                    f"  Исходный файл (src): '{src}'\n"
                    f"  Файл назначения (dst): '{dst}'\n"
                    f"  Ошибка: {e}",
                    exc_info=True,
                )
                # Пробрасываем дальше, чтобы увидеть стек в консоли/GUI
                raise

            wb = openpyxl.load_workbook(dst)
        
        # Обработка всех разделов
        sections_data = [
            ('доходы', form_data.get('доходы_data', [])),
            ('расходы', form_data.get('расходы_data', [])),
            ('источники_финансирования', form_data.get('источники_финансирования_data', [])),
            ('консолидируемые_расчеты', form_data.get('консолидируемые_расчеты_data', []))
        ]
        
        for section_type, data in sections_data:
            if data:
                if section_type == 'консолидируемые_расчеты':
                    self._process_consolidated_section_in_original_form(wb, data)
                else:
                    self._process_section_in_original_form(wb, data, section_type)
        
        # Проверка дефицита/профицита
        if calculated_deficit_proficit:
            self._validate_deficit_proficit(wb, calculated_deficit_proficit)
        
        wb.save(output_file_path)
        return output_file_path
    
    def _process_section_in_original_form(self, wb: openpyxl.Workbook, data: List[dict], section_name: str):
        """Обработка раздела в исходной форме
        
        Args:
            wb: Рабочая книга Excel
            data: Список данных раздела
            section_name: Название раздела
        """
        if not data:
            return
        
        sheet_name = self.constants.SECTION_SHEETS.get(section_name)
        if sheet_name not in wb.sheetnames:
            return
        
        ws = wb[sheet_name]
        budget_columns = self.constants.BUDGET_COLUMNS
        
        # Подготавливаем DataFrame для валидации:
        # используем уже имеющиеся расчетные значения, если они есть,
        # иначе берем оригинальные как расчетные (как в калькуляторе).
        df_rows: List[Dict[str, Any]] = []
        for item in data:
            row = dict(item)  # копия базовых полей (уровень, исходная_строка и т.п.)
            orig_approved = (item.get('утвержденный') or {}) or {}
            orig_executed = (item.get('исполненный') or {}) or {}
            for col in budget_columns:
                approved_val = orig_approved.get(col, 0)
                executed_val = orig_executed.get(col, 0)
                calc_approved = item.get(f'расчетный_утвержденный_{col}', approved_val)
                calc_executed = item.get(f'расчетный_исполненный_{col}', executed_val)
                row[f'утвержденный_{col}'] = approved_val
                row[f'исполненный_{col}'] = executed_val
                row[f'расчетный_утвержденный_{col}'] = calc_approved
                row[f'расчетный_исполненный_{col}'] = calc_executed
            df_rows.append(row)

        df = pd.DataFrame(df_rows)
        # Данные уже должны быть пересчитаны, поэтому просто применяем валидацию
        self._apply_validation_to_original_cells(ws, df, budget_columns, section_name)
    
    def _process_consolidated_section_in_original_form(self, wb: openpyxl.Workbook, data: List[dict]):
        """Обработка консолидируемых расчетов в исходной форме
        
        Args:
            wb: Рабочая книга Excel
            data: Список данных консолидированных расчетов
        """
        if not data:
            return
        
        sheet_name = self.constants.SECTION_SHEETS['консолидируемые_расчеты']
        if sheet_name not in wb.sheetnames:
            return
        
        ws = wb[sheet_name]
        consolidated_columns = self.constants.CONSOLIDATED_COLUMNS
        
        # Подготавливаем DataFrame для валидации:
        # используем уже имеющиеся расчетные значения, если они есть.
        df_rows: List[Dict[str, Any]] = []
        for item in data:
            row = dict(item)
            orig_receipts = (item.get('поступления') or {}) or {}
            for col in consolidated_columns:
                orig_val = orig_receipts.get(col, 0)
                calc_val = item.get(f'расчетный_поступления_{col}', orig_val)
                row[f'поступления_{col}'] = orig_val
                row[f'расчетный_поступления_{col}'] = calc_val
            df_rows.append(row)

        df = pd.DataFrame(df_rows)
        # Данные уже должны быть пересчитаны, поэтому просто применяем валидацию
        self._apply_consolidated_validation_to_original_cells(ws, df, consolidated_columns)
    
    def _apply_validation_to_original_cells(
        self, 
        ws, 
        df: pd.DataFrame, 
        budget_columns: List[str], 
        section_name: str
    ):
        """Применение проверки к исходным ячейкам
        
        Args:
            ws: Рабочий лист Excel
            df: DataFrame с данными (уже с расчетными значениями)
            budget_columns: Список бюджетных колонок
            section_name: Название раздела
        """
        for row_idx, (_, row_data) in enumerate(df.iterrows(), 0):
            # При экспорте с пересчетом обрабатываем все столбцы
            self._apply_row_validation(ws, row_data, budget_columns, [], section_name)
    
    def _apply_row_validation(
        self, 
        ws, 
        row_data: dict, 
        budget_columns: List[str], 
        zero_columns: List[int], 
        section_name: str
    ):
        """Применение проверки к строке
        
        Args:
            ws: Рабочий лист Excel
            row_data: Данные строки
            budget_columns: Список бюджетных колонок
            zero_columns: Список индексов нулевых колонок (не используется при экспорте)
            section_name: Название раздела
        """
        level = row_data['уровень']
        original_row = row_data['исходная_строка']
        mapping = self.constants.COLUMN_MAPPING[section_name]
        
        if level in StyleConstants.LEVEL_COLORS:
            self._fill_row_with_level_color(ws, original_row, level)
        
        for i, budget_col in enumerate(budget_columns):
            if i in zero_columns or (i + len(budget_columns)) in zero_columns:
                continue
            
            self._validate_budget_cells(ws, row_data, budget_col, i, mapping, original_row, level)
    
    def _fill_row_with_level_color(self, ws, row: int, level: int):
        """Закрашивание строки по уровню
        
        Args:
            ws: Рабочий лист Excel
            row: Номер строки
            level: Уровень строки
        """
        level_fill = PatternFill(
            start_color=StyleConstants.LEVEL_COLORS[level], 
            end_color=StyleConstants.LEVEL_COLORS[level], 
            fill_type="solid"
        )
        
        for col_idx in range(1, 37):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = level_fill
    
    def _validate_budget_cells(
        self, 
        ws, 
        row_data: dict, 
        budget_col: str, 
        col_index: int, 
        mapping: dict, 
        original_row: int, 
        level: int
    ):
        """Проверка бюджетных ячеек
        
        Args:
            ws: Рабочий лист Excel
            row_data: Данные строки
            budget_col: Название бюджетной колонки
            col_index: Индекс колонки
            mapping: Маппинг колонок
            original_row: Номер исходной строки
            level: Уровень строки
        """
        approved_col = mapping['утвержденный'][col_index]
        approved_col_idx = column_to_index(approved_col) + 1
        approved_cell = ws.cell(row=original_row, column=approved_col_idx)
        
        original_approved = row_data[f'утвержденный_{budget_col}']
        calculated_approved = row_data[f'расчетный_утвержденный_{budget_col}']
        
        if is_value_different(original_approved, calculated_approved) and level < 6:
            approved_cell.fill = StyleConstants.ERROR_FILL
            if self.show_error_values:
                try:
                    orig_val = float(original_approved or 0)
                except (TypeError, ValueError):
                    orig_val = 0.0
                try:
                    calc_val = float(calculated_approved or 0)
                except (TypeError, ValueError):
                    calc_val = 0.0
                approved_cell.value = f"{orig_val:.2f} ({calc_val:.2f})"
        
        executed_col = mapping['исполненный'][col_index]
        executed_col_idx = column_to_index(executed_col) + 1
        executed_cell = ws.cell(row=original_row, column=executed_col_idx)
        
        original_executed = row_data[f'исполненный_{budget_col}']
        calculated_executed = row_data[f'расчетный_исполненный_{budget_col}']
        
        if is_value_different(original_executed, calculated_executed) and level < 6:
            executed_cell.fill = StyleConstants.ERROR_FILL
            if self.show_error_values:
                try:
                    orig_val = float(original_executed or 0)
                except (TypeError, ValueError):
                    orig_val = 0.0
                try:
                    calc_val = float(calculated_executed or 0)
                except (TypeError, ValueError):
                    calc_val = 0.0
                executed_cell.value = f"{orig_val:.2f} ({calc_val:.2f})"
    
    def _apply_consolidated_validation_to_original_cells(
        self, 
        ws, 
        df: pd.DataFrame, 
        consolidated_columns: List[str]
    ):
        """Применение проверки к ячейкам консолидируемых расчетов
        
        Args:
            ws: Рабочий лист Excel
            df: DataFrame с данными
            consolidated_columns: Список колонок консолидированных расчетов
        """
        for row_idx, (_, row_data) in enumerate(df.iterrows(), 0):
            self._apply_consolidated_row_validation(ws, row_data, consolidated_columns)
    
    def _apply_consolidated_row_validation(self, ws, row_data: dict, consolidated_columns: List[str]):
        """Применение проверки к строке консолидируемых расчетов
        
        Args:
            ws: Рабочий лист Excel
            row_data: Данные строки
            consolidated_columns: Список колонок консолидированных расчетов
        """
        level = row_data['уровень']
        original_row = row_data['исходная_строка']
        mapping = self.constants.COLUMN_MAPPING['консолидируемые_расчеты']
        
        if level in StyleConstants.LEVEL_COLORS:
            self._fill_consolidated_row_with_level_color(ws, original_row, level)
        
        budget_columns = consolidated_columns[:-1]
        for i, budget_col in enumerate(budget_columns):
            self._validate_consolidated_cells(ws, row_data, budget_col, i, mapping, original_row, level)
        
        self._validate_sum_column(ws, row_data, consolidated_columns, mapping, original_row, level)
    
    def _fill_consolidated_row_with_level_color(self, ws, row: int, level: int):
        """Закрашивание строки консолидируемых расчетов
        
        Args:
            ws: Рабочий лист Excel
            row: Номер строки
            level: Уровень строки
        """
        level_fill = PatternFill(
            start_color=StyleConstants.LEVEL_COLORS[level], 
            end_color=StyleConstants.LEVEL_COLORS[level], 
            fill_type="solid"
        )
        
        for col_idx in range(2, 15):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = level_fill
    
    def _validate_consolidated_cells(
        self, 
        ws, 
        row_data: dict, 
        budget_col: str, 
        col_index: int, 
        mapping: dict, 
        original_row: int, 
        level: int
    ):
        """Проверка ячеек консолидируемых расчетов
        
        Args:
            ws: Рабочий лист Excel
            row_data: Данные строки
            budget_col: Название бюджетной колонки
            col_index: Индекс колонки
            mapping: Маппинг колонок
            original_row: Номер исходной строки
            level: Уровень строки
        """
        receipt_col = mapping['поступления'][col_index]
        receipt_col_idx = column_to_index(receipt_col) + 1
        receipt_cell = ws.cell(row=original_row, column=receipt_col_idx)
        
        original_receipt = row_data[f'поступления_{budget_col}']
        calculated_receipt = row_data[f'расчетный_поступления_{budget_col}']
        
        if original_receipt == 'x':
            return
        
        if is_value_different(original_receipt, calculated_receipt) and level < 2:
            receipt_cell.fill = StyleConstants.ERROR_FILL
            if self.show_error_values:
                try:
                    orig_val = float(original_receipt or 0)
                except (TypeError, ValueError):
                    orig_val = 0.0
                try:
                    calc_val = float(calculated_receipt or 0)
                except (TypeError, ValueError):
                    calc_val = 0.0
                receipt_cell.value = f"{orig_val:.2f} ({calc_val:.2f})"
    
    def _validate_sum_column(
        self, 
        ws, 
        row_data: dict, 
        consolidated_columns: List[str], 
        mapping: dict, 
        original_row: int, 
        level: int
    ):
        """Проверка столбца 'ИТОГО'
        
        Args:
            ws: Рабочий лист Excel
            row_data: Данные строки
            consolidated_columns: Список колонок консолидированных расчетов
            mapping: Маппинг колонок
            original_row: Номер исходной строки
            level: Уровень строки
        """
        sum_col_index = len(consolidated_columns) - 1
        sum_col = mapping['поступления'][sum_col_index]
        sum_col_idx = column_to_index(sum_col) + 1
        sum_cell = ws.cell(row=original_row, column=sum_col_idx)
        
        # Последний столбец в CONSOLIDATED_COLUMNS содержит агрегированное значение (ИТОГО)
        sum_column_name = consolidated_columns[-1]
        original_sum = row_data.get(f'поступления_{sum_column_name}')
        
        # Используем уже рассчитанное значение, если оно есть
        calculated_sum = row_data.get(f'расчетный_поступления_{sum_column_name}')
        
        # Если расчетное значение не найдено, пересчитываем сумму из бюджетных столбцов
        if calculated_sum is None:
            calculated_sum = 0.0
            budget_columns = consolidated_columns[:-1]
            
            for budget_col in budget_columns:
                calculated_value = row_data.get(f'расчетный_поступления_{budget_col}')
                if calculated_value is None:
                    calculated_value = row_data.get(f'поступления_{budget_col}', 0)
                
                if calculated_value != 'x' and calculated_value is not None:
                    calculated_sum += calculated_value
        
        if original_sum == 'x':
            return
        
        if is_value_different(original_sum, calculated_sum) and level < 2:
            sum_cell.fill = StyleConstants.ERROR_FILL
            if self.show_error_values:
                try:
                    orig_val = float(original_sum or 0)
                except (TypeError, ValueError):
                    orig_val = 0.0
                try:
                    calc_val = float(calculated_sum or 0)
                except (TypeError, ValueError):
                    calc_val = 0.0
                sum_cell.value = f"{orig_val:.2f} ({calc_val:.2f})"
    
    def _validate_deficit_proficit(
        self, 
        wb: openpyxl.Workbook, 
        calculated_deficit_proficit: Dict[str, Dict[str, float]]
    ):
        """Проверка дефицита/профицита
        
        Args:
            wb: Рабочая книга Excel
            calculated_deficit_proficit: Рассчитанный дефицит/профицит
        """
        if not calculated_deficit_proficit:
            return
        
        sheet_name = self.constants.SECTION_SHEETS['расходы']
        if sheet_name not in wb.sheetnames:
            return
        
        ws = wb[sheet_name]
        budget_columns = self.constants.BUDGET_COLUMNS
        mapping = self.constants.COLUMN_MAPPING['расходы']
        
        target_row = None
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            code_cell = ws.cell(row=row, column=2).value
            
            # Ищем строку, где одновременно есть "результат исполнения бюджета" и код 450
            if (cell_value and 'результат исполнения бюджета' in str(cell_value).lower() and
                code_cell and '450' in str(code_cell)):
                target_row = row
                break
        
        if not target_row:
            return
        
        for i, budget_col in enumerate(budget_columns):
            approved_col = mapping['утвержденный'][i]
            approved_col_idx = column_to_index(approved_col) + 1
            approved_cell = ws.cell(row=target_row, column=approved_col_idx)
            
            # Используем метод для извлечения оригинального значения из ячейки
            original_approved = extract_original_value_from_cell(approved_cell.value)
            calculated_approved = calculated_deficit_proficit['утвержденный'][budget_col]
            
            if is_value_different(original_approved, calculated_approved):
                approved_cell.fill = StyleConstants.ERROR_FILL
                if self.show_error_values:
                    try:
                        orig_val = float(original_approved or 0)
                    except (TypeError, ValueError):
                        orig_val = 0.0
                    try:
                        calc_val = float(calculated_approved or 0)
                    except (TypeError, ValueError):
                        calc_val = 0.0
                    approved_cell.value = f"{orig_val:.2f} ({calc_val:.2f})"
            
            executed_col = mapping['исполненный'][i]
            executed_col_idx = column_to_index(executed_col) + 1
            executed_cell = ws.cell(row=target_row, column=executed_col_idx)
            
            # Используем метод для извлечения оригинального значения из ячейки
            original_executed = extract_original_value_from_cell(executed_cell.value)
            calculated_executed = calculated_deficit_proficit['исполненный'][budget_col]
            
            if is_value_different(original_executed, calculated_executed):
                executed_cell.fill = StyleConstants.ERROR_FILL
                if self.show_error_values:
                    try:
                        orig_val = float(original_executed or 0)
                    except (TypeError, ValueError):
                        orig_val = 0.0
                    try:
                        calc_val = float(calculated_executed or 0)
                    except (TypeError, ValueError):
                        calc_val = 0.0
                    executed_cell.value = f"{orig_val:.2f} ({calc_val:.2f})"
    
    def _hide_zero_columns(self, ws, section_name: str, zero_columns: List[int]):
        """Скрытие нулевых столбцов
        
        Args:
            ws: Рабочий лист Excel
            section_name: Название раздела
            zero_columns: Список индексов нулевых колонок
        """
        mapping = self.constants.COLUMN_MAPPING[section_name]
        budget_columns = self.constants.BUDGET_COLUMNS
        
        for col_index in zero_columns:
            if col_index < len(budget_columns):
                col_letter = mapping['утвержденный'][col_index]
            else:
                col_letter = mapping['исполненный'][col_index - len(budget_columns)]
            
            col_idx = column_to_index(col_letter) + 1
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].hidden = True
