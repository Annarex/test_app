import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
import shutil
from typing import Dict, List, Any, Optional
from datetime import datetime
from .base_models import BaseFormModel, FormType
from logger import logger
from models.constants.form_0503317_constants import Form0503317Constants, StyleConstants
from models.parsers.form_0503317_parser import Form0503317Parser
from models.calculators.form_0503317_calculator import Form0503317Calculator
from models.exporters.form_0503317_exporter import Form0503317Exporter
from models.utils.form_utils import (
    column_to_index, get_cell_value, get_numeric_value,
    clean_dbk_code, format_classification_code, extract_original_value_from_cell
)
from utils.numeric_utils import is_value_different

class Form0503317(BaseFormModel):
    """Модель для формы 0503317"""
    
    def __init__(self, revision: str = "1.0", column_mapping: Optional[dict] = None):
        super().__init__(FormType.FORM_0503317, revision)
        self.constants = Form0503317Constants()
        if column_mapping:
            # Позволяем переопределять mapping колонок из справочника типов форм
            self.constants.COLUMN_MAPPING = column_mapping
        
        # Инициализация компонентов через композицию
        self.parser = Form0503317Parser(self.constants)
        self.calculator = Form0503317Calculator(self.constants)
        self.exporter = Form0503317Exporter(self.constants, self.calculator)
        
        self.meta_info = {}
        self.доходы_data = []
        self.расходы_data = []
        self.источники_финансирования_data = []
        self.консолидируемые_расчеты_data = []
        self.reference_data_доходы = None
        self.reference_data_источники = None
        self.zero_columns = {}
        self.show_error_values = True
        self.доходы_всего = None
        self.расходы_всего = None
        self.calculated_deficit_proficit = None  # Переименовано из результат_исполнения_data для ясности
    
    def get_form_constants(self):
        return self.constants
    
    def parse_excel(self, file_path: str, reference_data_доходы: pd.DataFrame = None, reference_data_источники: pd.DataFrame = None) -> Dict[str, Any]:
        """Парсинг Excel файла формы 0503317"""
        # Сбрасываем предыдущее состояние, чтобы при повторной загрузке формы
        # данные не дублировались
        self.meta_info = {}
        self.доходы_data = []
        self.расходы_data = []
        self.источники_финансирования_data = []
        self.консолидируемые_расчеты_data = []
        self.zero_columns = {}
        self.доходы_всего = None
        self.расходы_всего = None
        self.calculated_deficit_proficit = None

        self.reference_data_доходы = reference_data_доходы
        self.reference_data_источники = reference_data_источники
        
        # Используем парсер для извлечения данных
        parsed_data = self.parser.parse_excel(file_path, reference_data_доходы, reference_data_источники)
        
        # Обновляем внутреннее состояние
        self.meta_info = parsed_data.get('meta_info', {})
        self.доходы_data = parsed_data.get('доходы_data', [])
        self.расходы_data = parsed_data.get('расходы_data', [])
        self.источники_финансирования_data = parsed_data.get('источники_финансирования_data', [])
        self.консолидируемые_расчеты_data = parsed_data.get('консолидируемые_расчеты_data', [])
        self.zero_columns = parsed_data.get('zero_columns', {})
        
        # Рассчитываем дефицит/профицит
        self.calculated_deficit_proficit = self.calculator.calculate_deficit_proficit(
            self.доходы_data,
            self.расходы_data
        )
        
        return {
            'meta_info': self.meta_info,
            'доходы_data': self.доходы_data,
            'расходы_data': self.расходы_data,
            'источники_финансирования_data': self.источники_финансирования_data,
            'консолидируемые_расчеты_data': self.консолидируемые_расчеты_data,
            'calculated_deficit_proficit': self.calculated_deficit_proficit
        }
    
    def calculate_sums(self) -> Dict[str, Any]:
        """Расчет агрегированных сумм"""
        # Формируем данные для калькулятора
        form_data = {
            'доходы_data': self.доходы_data,
            'расходы_data': self.расходы_data,
            'источники_финансирования_data': self.источники_финансирования_data,
            'консолидируемые_расчеты_data': self.консолидируемые_расчеты_data
        }
        
        # Используем калькулятор для расчета
        calculated_data = self.calculator.calculate_sums(form_data)
        
        # Обновляем внутреннее состояние
        self.доходы_data = calculated_data.get('доходы_data', self.доходы_data)
        self.расходы_data = calculated_data.get('расходы_data', self.расходы_data)
        self.источники_финансирования_data = calculated_data.get('источники_финансирования_data', self.источники_финансирования_data)
        self.консолидируемые_расчеты_data = calculated_data.get('консолидируемые_расчеты_data', self.консолидируемые_расчеты_data)
        
        # Пересчитываем дефицит/профицит
        self.calculated_deficit_proficit = self.calculator.calculate_deficit_proficit(
            self.доходы_data,
            self.расходы_data
        )
        
        # calculate_sums возвращает только данные разделов для пересчета
        # meta_info и calculated_deficit_proficit не должны возвращаться здесь,
        # они уже есть в форме и сохраняются отдельно при сохранении ревизии
        return calculated_data

    def recalculate_levels_with_references(
        self,
        form_data: Dict[str, Any],
        reference_data_доходы: Optional[pd.DataFrame] = None,
        reference_data_источники: Optional[pd.DataFrame] = None
    ) -> Dict[str, Any]:
        """
        Пересчет уровней строк на основе актуальных справочников.
        Используется при повторной загрузке уже сохраненного проекта.
        """
        self.reference_data_доходы = reference_data_доходы
        self.reference_data_источники = reference_data_источники
        
        # Обновляем справочники в парсере
        self.parser.reference_data_доходы = reference_data_доходы
        self.parser.reference_data_источники = reference_data_источники

        # Доходы
        доходы_data = form_data.get('доходы_data', [])
        for item in доходы_data:
            code = item.get('код_классификации', '')
            name = item.get('наименование_показателя', '')
            item['уровень'] = self.parser._determine_level(code, 'доходы', name)

        # Источники финансирования
        источники_data = form_data.get('источники_финансирования_data', [])
        for item in источники_data:
            code = item.get('код_классификации', '')
            name = item.get('наименование_показателя', '')
            item['уровень'] = self.parser._determine_level(code, 'источники_финансирования', name)

        # Расходы зависят только от кода, справочники не нужны
        расходы_data = form_data.get('расходы_data', [])
        for item in расходы_data:
            code = item.get('код_классификации', '')
            item['уровень'] = self.parser._determine_expenditure_level(code)

        form_data['доходы_data'] = доходы_data
        form_data['расходы_data'] = расходы_data
        form_data['источники_финансирования_data'] = источники_data
        return form_data

    def load_saved_data(self, form_data: Dict[str, Any]):
        """
        Инициализация внутренних структур формы из сохраненных данных проекта.
        Требуется при повторной загрузке проекта, чтобы экспорт/проверка работали корректно.
        """
        self.meta_info = form_data.get('meta_info', {})
        self.доходы_data = form_data.get('доходы_data', []) or []
        self.расходы_data = form_data.get('расходы_data', []) or []
        self.источники_финансирования_data = form_data.get('источники_финансирования_data', []) or []
        self.консолидируемые_расчеты_data = form_data.get('консолидируемые_расчеты_data', []) or []

        # Восстанавливаем вспомогательные структуры
        self.zero_columns = {}
        if self.доходы_data:
            total_row = next((item for item in self.доходы_data if 'всего' in item.get('наименование_показателя', '').lower()), None)
            if total_row:
                self.zero_columns['доходы'] = self.parser._get_zero_columns(total_row, self.constants.BUDGET_COLUMNS)

        if self.расходы_data:
            total_row = next((item for item in self.расходы_data if 'всего' in item.get('наименование_показателя', '').lower()), None)
            if total_row:
                self.zero_columns['расходы'] = self.parser._get_zero_columns(total_row, self.constants.BUDGET_COLUMNS)

        if self.источники_финансирования_data:
            total_row = next((item for item in self.источники_финансирования_data if 'всего' in item.get('наименование_показателя', '').lower()), None)
            if total_row:
                self.zero_columns['источники_финансирования'] = self.parser._get_zero_columns(total_row, self.constants.BUDGET_COLUMNS)

        # Источники требуют корректировки уровней
        self.parser._recalculate_sources_levels(self.источники_финансирования_data)

        # Пересчитываем итоговые значения для дефицита/профицита, если их нет
        if not self.calculated_deficit_proficit:
            self.calculated_deficit_proficit = self.calculator.calculate_deficit_proficit(
                self.доходы_data,
                self.расходы_data
            )
    
    def validate_data(self) -> List[Dict[str, Any]]:
        """Валидация данных"""
        errors = []
        
        # Проверка агрегации сумм
        for section, data in [
            ('Доходы', self.доходы_data),
            ('Расходы', self.расходы_data),
            ('Источники финансирования', self.источники_финансирования_data),
            ('Консолидируемые расчеты', self.консолидируемые_расчеты_data)
        ]:
            if data:
                section_errors = self._validate_section_aggregation(section, data)
                errors.extend(section_errors)
        
        return errors
    
    def export_validation(self, original_file_path: str, output_file_path: str) -> str:
        """Экспорт формы с проверкой"""
        # Пересчитываем дефицит/профицит по текущим данным формы перед проверкой
        if not self.calculated_deficit_proficit:
            self.calculated_deficit_proficit = self.calculator.calculate_deficit_proficit(
                self.доходы_data,
                self.расходы_data
            )
        
        # Формируем данные для экспортёра
        form_data = {
            'доходы_data': self.доходы_data,
            'расходы_data': self.расходы_data,
            'источники_финансирования_data': self.источники_финансирования_data,
            'консолидируемые_расчеты_data': self.консолидируемые_расчеты_data
        }
        
        # Устанавливаем настройку отображения ошибок
        self.exporter.show_error_values = self.show_error_values
        
        # Используем экспортёр для валидации и экспорта
        return self.exporter.export_validation(
            original_file_path,
            output_file_path,
            form_data,
            self.calculated_deficit_proficit
        )
    
    # Вспомогательные методы
    def _extract_metadata(self, sheet: pd.DataFrame):
        """Извлечение метаданных (deprecated - используйте parser._extract_metadata)"""
        self.meta_info = {
            'Наименование формы': get_cell_value(sheet, 2, 1) + ' ' + get_cell_value(sheet, 3, 1),
            'Наименование финансового органа': get_cell_value(sheet, 6, 3),
            'Наименование бюджета': get_cell_value(sheet, 7, 3),
            'Периодичность': get_cell_value(sheet, 8, 17),
            'Форма по ОКУД': get_cell_value(sheet, 4, 17),
            'Дата': get_cell_value(sheet, 5, 17),
            'код ОКПО': get_cell_value(sheet, 6, 17),
            'код ОКТМО': get_cell_value(sheet, 7, 17),
            'код ОКЕИ': get_cell_value(sheet, 9, 17),
        }
    
    def _find_section_start(self, sheet: pd.DataFrame, section_type: str, search_column: int = 0) -> int:
        """Поиск начала раздела (deprecated - используйте parser._find_section_start)"""
        return self.parser._find_section_start(sheet, section_type, search_column)
    
    def _extract_consolidated_data(self, sheet: pd.DataFrame):
        """Извлечение данных консолидируемых расчетов (deprecated - используйте parser)"""
        # Метод больше не используется, так как parse_excel использует parser напрямую
        pass
    
    def _extract_consolidated_table_data(self, sheet: pd.DataFrame, start_row: int):
        """Извлечение данных таблицы консолидируемых расчетов (deprecated - используйте parser)"""
        # Метод больше не используется
        return []
    
    def _extract_consolidated_part_data(self, sheet: pd.DataFrame, start_row: int, mapping: dict, consolidated_columns: list) -> list:
        """Извлечение данных части таблицы консолидируемых расчетов (deprecated - используйте parser)"""
        # Метод больше не используется
        return []
    
    def _extract_consolidated_row_data(self, sheet: pd.DataFrame, row_idx: int, mapping: dict, consolidated_columns: list) -> dict:
        """Извлечение данных строки консолидируемых расчетов (deprecated - используйте parser)"""
        # Метод больше не используется
        return None
    
    def _determine_consolidated_level(self, code: str) -> int:
        """Определение уровня для консолидируемых расчетов (deprecated - используйте parser._determine_consolidated_level)"""
        return self.parser._determine_consolidated_level(code)
    
    def _extract_section_data(self, sheet: pd.DataFrame, section_type: str):
        """Извлечение данных раздела (deprecated - используйте parser)"""
        # Метод больше не используется, так как parse_excel использует parser напрямую
        pass
    
    def _extract_table_data(self, sheet: pd.DataFrame, header_row: int, section_type: str):
        """Извлечение табличных данных (deprecated - используйте parser)"""
        # Метод больше не используется
        pass
    
    def _extract_row_data(self, sheet: pd.DataFrame, row_idx: int, mapping: dict, budget_columns: list, section_type: str) -> dict:
        """Извлечение данных строки (deprecated - используйте parser)"""
        # Метод больше не используется
        return None
    
    def _extract_budget_data(self, sheet: pd.DataFrame, row_idx: int, columns: list, budget_columns: list) -> dict:
        """Извлечение данных по бюджету (deprecated - используйте parser)"""
        # Метод больше не используется
        return {}
    
    def _determine_level(self, classification_code: str, section_type: str, name: str = "") -> int:
        """Определение уровня строки (deprecated - используйте parser._determine_level)"""
        return self.parser._determine_level(classification_code, section_type, name)
    
    def _determine_expenditure_level(self, code: str) -> int:
        """Определение уровня для расходов (deprecated - используйте parser._determine_expenditure_level)"""
        return self.parser._determine_expenditure_level(code)
    
    def _get_level_from_reference(self, classification_code: str, section_type: str) -> int:
        """Получение уровня из справочника (deprecated - используйте parser._get_level_from_reference)"""
        return self.parser._get_level_from_reference(classification_code, section_type)
    
    def _is_total_row(self, row_data: dict, section_type: str) -> bool:
        """Проверка, является ли строка итоговой (deprecated - используйте parser._is_total_row)"""
        return self.parser._is_total_row(row_data, section_type)
    
    def _get_zero_columns(self, total_row_data: dict, budget_columns: list) -> list:
        """Получение нулевых столбцов (deprecated - используйте parser._get_zero_columns)"""
        return self.parser._get_zero_columns(total_row_data, budget_columns)
    
    def _recalculate_sources_levels(self):
        """Пересчет уровней для источников финансирования (deprecated - используйте parser._recalculate_sources_levels)"""
        self.parser._recalculate_sources_levels(self.источники_финансирования_data)
    
    def _find_total_row(self, data: list, pattern: str) -> dict:
        """Поиск итоговой строки по паттерну (deprecated - используйте calculator._find_total_row)"""
        return self.calculator._find_total_row(data, pattern)
    
    def _calculate_deficit_proficit_from_original(self, original_доходы_data: list = None, original_расходы_data: list = None):
        """Расчет дефицита/профицита из исходных данных (deprecated - используйте calculator.calculate_deficit_proficit)"""
        доходы_data = original_доходы_data if original_доходы_data is not None else self.доходы_data
        расходы_data = original_расходы_data if original_расходы_data is not None else self.расходы_data
        
        result = self.calculator.calculate_deficit_proficit(доходы_data, расходы_data)
        if result:
            self.calculated_deficit_proficit = result
        return result
    
    def _calculate_deficit_proficit(self):
        """Расчет дефицита/профицита (использует текущие данные формы)
        
        Deprecated: используйте calculator.calculate_deficit_proficit() напрямую
        """
        self.calculated_deficit_proficit = self.calculator.calculate_deficit_proficit(
            self.доходы_data,
            self.расходы_data
        )
        return self.calculated_deficit_proficit
    
    def _prepare_dataframe_for_calculation(self, data: list, budget_columns: list) -> pd.DataFrame:
        """Подготовка DataFrame для вычислений (deprecated - используйте calculator._prepare_dataframe_for_calculation)"""
        return self.calculator._prepare_dataframe_for_calculation(data, budget_columns)
    
    def _prepare_consolidated_dataframe_for_calculation(self, data: list, consolidated_columns: list) -> pd.DataFrame:
        """Подготовка DataFrame для консолидируемых расчетов (deprecated - используйте calculator._prepare_consolidated_dataframe_for_calculation)"""
        return self.calculator._prepare_consolidated_dataframe_for_calculation(data, consolidated_columns)
    
    def _calculate_budget_sums(self, df: pd.DataFrame, budget_columns: list) -> pd.DataFrame:
        """Расчет бюджетных сумм (deprecated - используйте calculator._calculate_budget_sums)"""
        return self.calculator._calculate_budget_sums(df, budget_columns)
    
    def _calculate_standard_sums(self, df: pd.DataFrame, budget_columns: list) -> pd.DataFrame:
        """Стандартный расчет сумм (deprecated - используйте calculator._calculate_standard_sums)"""
        return self.calculator._calculate_standard_sums(df, budget_columns)
    
    def _calculate_sources_sums(self, df: pd.DataFrame, budget_columns: list) -> pd.DataFrame:
        """Расчет сумм для источников финансирования (deprecated - используйте calculator._calculate_sources_sums)"""
        return self.calculator._calculate_sources_sums(df, budget_columns)
    
    def _calculate_consolidated_sums(self, df: pd.DataFrame) -> pd.DataFrame:
        """Расчет сумм для консолидируемых расчетов (deprecated - используйте calculator._calculate_consolidated_sums)"""
        return self.calculator._calculate_consolidated_sums(df)
    
    def _sum_level1_for_level0(self, df: pd.DataFrame, parent_idx: int, column: str):
        """Суммирование значений уровня 1 для уровня 0 (deprecated - используйте calculator._sum_level1_for_level0)"""
        return self.calculator._sum_level1_for_level0(df, parent_idx, column)
    
    def _sum_consolidated_children(self, df: pd.DataFrame, parent_idx: int, start_idx: int, end_idx: int, column: str, current_level: int):
        """Суммирование дочерних элементов для консолидируемых расчетов (deprecated - используйте calculator._sum_consolidated_children)"""
        return self.calculator._sum_consolidated_children(df, parent_idx, start_idx, end_idx, column, current_level)
    
    def _find_child_boundaries(self, df: pd.DataFrame, current_idx: int, current_level: int) -> tuple:
        """Поиск границ дочерних элементов (deprecated - используйте calculator._find_child_boundaries)"""
        return self.calculator._find_child_boundaries(df, current_idx, current_level)
    
    def _sum_children_for_budget_column(self, df: pd.DataFrame, parent_idx: int, start_idx: int, end_idx: int, budget_col: str, current_level: int):
        """Суммирование дочерних элементов для бюджетной колонки (deprecated - используйте calculator._sum_children_for_budget_column)"""
        return self.calculator._sum_children_for_budget_column(df, parent_idx, start_idx, end_idx, budget_col, current_level)
    
    def _validate_section_aggregation(self, section_name: str, data: List[Dict]) -> List[Dict[str, Any]]:
        """Валидация агрегации раздела"""
        errors = []
        
        for item in data:
            level = item.get('уровень', 0)
            if level < 6:  # Проверяем только агрегирующие уровни
                # Проверка бюджетных колонок
                for budget_col in self.constants.BUDGET_COLUMNS:
                    original_approved = item.get('утвержденный', {}).get(budget_col, 0)
                    calculated_approved = item.get('расчетный_утвержденный_{budget_col}', original_approved)
                    
                    if abs(original_approved - calculated_approved) > 0.00001:
                        errors.append({
                            'section': section_name,
                            'row': item.get('наименование_показателя', ''),
                            'level': level,
                            'column': f'утвержденный_{budget_col}',
                            'original': original_approved,
                            'calculated': calculated_approved
                        })
        
        return errors
    
    def _process_consolidated_section_in_original_form(self, wb: openpyxl.Workbook, data: list):
        """Обработка консолидируемых расчетов в исходной форме (deprecated - используйте exporter)"""
        # Метод больше не используется, так как export_validation использует exporter напрямую
        pass
    
    def _process_section_in_original_form(self, wb: openpyxl.Workbook, data: list, section_name: str):
        """Обработка раздела в исходной форме (deprecated - используйте exporter)"""
        # Метод больше не используется, так как export_validation использует exporter напрямую
        pass
    
    def _apply_consolidated_validation_to_original_cells(self, ws, df: pd.DataFrame, consolidated_columns: list):
        """Применение проверки к ячейкам консолидируемых расчетов"""
        for row_idx, (_, row_data) in enumerate(df.iterrows(), 0):
            self._apply_consolidated_row_validation(ws, row_data, consolidated_columns)
    
    def _apply_consolidated_row_validation(self, ws, row_data: dict, consolidated_columns: list):
        """Применение проверки к строке консолидируемых расчетов"""
        level = row_data['уровень']
        original_row = row_data['исходная_строка']
        mapping = self.constants.COLUMN_MAPPING['консолидируемые_расчеты']
        
        if level in StyleConstants.LEVEL_COLORS:
            self._fill_consolidated_row_with_level_color(ws, original_row, level)
        
        budget_columns = consolidated_columns[:-1]
        for i, budget_col in enumerate(budget_columns):
            self._validate_consolidated_cells(ws, row_data, budget_col, i, mapping, original_row, level)
        
        self._validate_sum_column(ws, row_data, consolidated_columns, mapping, original_row, level)
    
    def _fill_consolidated_row_with_level_color(self, ws, row: int, level: int):
        """Закрашивание строки консолидируемых расчетов"""
        level_fill = PatternFill(
            start_color=StyleConstants.LEVEL_COLORS[level], 
            end_color=StyleConstants.LEVEL_COLORS[level], 
            fill_type="solid"
        )
        
        for col_idx in range(2, 15):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = level_fill
    
    def _validate_consolidated_cells(self, ws, row_data: dict, budget_col: str, col_index: int, mapping: dict, original_row: int, level: int):
        """Проверка ячеек консолидируемых расчетов"""
        receipt_col = mapping['поступления'][col_index]
        receipt_col_idx = column_to_index(receipt_col) + 1
        receipt_cell = ws.cell(row=original_row, column=receipt_col_idx)
        
        original_receipt = row_data[f'поступления_{budget_col}']
        calculated_receipt = row_data[f'расчетный_поступления_{budget_col}']
        
        if original_receipt == 'x':
            return
        
        if is_value_different(original_receipt, calculated_receipt) and level < 2:
            receipt_cell.fill = StyleConstants.ERROR_FILL
            if self.show_error_values:
                try:
                    orig_val = float(original_receipt or 0)
                except (TypeError, ValueError):
                    orig_val = 0.0
                try:
                    calc_val = float(calculated_receipt or 0)
                except (TypeError, ValueError):
                    calc_val = 0.0
                receipt_cell.value = f"{orig_val:.2f} ({calc_val:.2f})"
    
    def _validate_sum_column(self, ws, row_data: dict, consolidated_columns: list, mapping: dict, original_row: int, level: int):
        """Проверка столбца 'ИТОГО'"""
        sum_col_index = len(consolidated_columns) - 1
        sum_col = mapping['поступления'][sum_col_index]
        sum_col_idx = column_to_index(sum_col) + 1
        sum_cell = ws.cell(row=original_row, column=sum_col_idx)
        
        # Последний столбец в CONSOLIDATED_COLUMNS содержит агрегированное значение (ИТОГО)
        sum_column_name = consolidated_columns[-1]
        original_sum = row_data.get(f'поступления_{sum_column_name}')
        
        # Используем уже рассчитанное значение, если оно есть
        calculated_sum = row_data.get(f'расчетный_поступления_{sum_column_name}')
        
        # Если расчетное значение не найдено, пересчитываем сумму из бюджетных столбцов
        if calculated_sum is None:
            calculated_sum = 0.0
            budget_columns = consolidated_columns[:-1]
            
            for budget_col in budget_columns:
                calculated_value = row_data.get(f'расчетный_поступления_{budget_col}')
                if calculated_value is None:
                    calculated_value = row_data.get(f'поступления_{budget_col}', 0)
                
                if calculated_value != 'x' and calculated_value is not None:
                    calculated_sum += calculated_value
        
        if original_sum == 'x':
            return
        
        if is_value_different(original_sum, calculated_sum) and level < 2:
            sum_cell.fill = StyleConstants.ERROR_FILL
            if self.show_error_values:
                try:
                    orig_val = float(original_sum or 0)
                except (TypeError, ValueError):
                    orig_val = 0.0
                try:
                    calc_val = float(calculated_sum or 0)
                except (TypeError, ValueError):
                    calc_val = 0.0
                sum_cell.value = f"{orig_val:.2f} ({calc_val:.2f})"
    
    def _validate_deficit_proficit(self, wb: openpyxl.Workbook):
        """Проверка дефицита/профицита"""
        if not self.calculated_deficit_proficit:
            return
        
        sheet_name = self.constants.SECTION_SHEETS['расходы']
        if sheet_name not in wb.sheetnames:
            return
        
        ws = wb[sheet_name]
        budget_columns = self.constants.BUDGET_COLUMNS
        mapping = self.constants.COLUMN_MAPPING['расходы']
        
        target_row = None
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            code_cell = ws.cell(row=row, column=2).value
            
            # Ищем строку, где одновременно есть "результат исполнения бюджета" и код 450
            if (cell_value and 'результат исполнения бюджета' in str(cell_value).lower() and
                code_cell and '450' in str(code_cell)):
                target_row = row
                break
        
        if not target_row:
            return
        
        for i, budget_col in enumerate(budget_columns):
            approved_col = mapping['утвержденный'][i]
            approved_col_idx = column_to_index(approved_col) + 1
            approved_cell = ws.cell(row=target_row, column=approved_col_idx)
            
            # Используем метод для извлечения оригинального значения из ячейки
            # (обрабатывает случаи, когда в ячейке уже есть формат "orig (calc)")
            original_approved = extract_original_value_from_cell(approved_cell.value)
            calculated_approved = self.calculated_deficit_proficit['утвержденный'][budget_col]
            
            if is_value_different(original_approved, calculated_approved):
                approved_cell.fill = StyleConstants.ERROR_FILL
                if self.show_error_values:
                    try:
                        orig_val = float(original_approved or 0)
                    except (TypeError, ValueError):
                        orig_val = 0.0
                    try:
                        calc_val = float(calculated_approved or 0)
                    except (TypeError, ValueError):
                        calc_val = 0.0
                    approved_cell.value = f"{orig_val:.2f} ({calc_val:.2f})"
            
            executed_col = mapping['исполненный'][i]
            executed_col_idx = column_to_index(executed_col) + 1
            executed_cell = ws.cell(row=target_row, column=executed_col_idx)
            
            # Используем метод для извлечения оригинального значения из ячейки
            # (обрабатывает случаи, когда в ячейке уже есть формат "orig (calc)")
            original_executed = extract_original_value_from_cell(executed_cell.value)
            calculated_executed = self.calculated_deficit_proficit['исполненный'][budget_col]
            
            if is_value_different(original_executed, calculated_executed):
                executed_cell.fill = StyleConstants.ERROR_FILL
                if self.show_error_values:
                    try:
                        orig_val = float(original_executed or 0)
                    except (TypeError, ValueError):
                        orig_val = 0.0
                    try:
                        calc_val = float(calculated_executed or 0)
                    except (TypeError, ValueError):
                        calc_val = 0.0
                    executed_cell.value = f"{orig_val:.2f} ({calc_val:.2f})"
    
    def _apply_validation_to_original_cells(self, ws, df: pd.DataFrame, budget_columns: list, section_name: str):
        """Применение проверки к исходным ячейкам"""
        zero_columns = self.zero_columns.get(section_name, [])
        
        for row_idx, (_, row_data) in enumerate(df.iterrows(), 0):
            # При экспорте с пересчетом обрабатываем все столбцы, включая нулевые
            # Передаем пустой список нулевых столбцов, чтобы не пропускать их
            self._apply_row_validation(ws, row_data, budget_columns, [], section_name)
        
        # При экспорте с пересчетом НЕ скрываем нулевые столбцы - выводим все столбцы
        # self._hide_zero_columns(ws, section_name, zero_columns)
    
    def _apply_row_validation(self, ws, row_data: dict, budget_columns: list, zero_columns: list, section_name: str):
        """Применение проверки к строке"""
        level = row_data['уровень']
        original_row = row_data['исходная_строка']
        mapping = self.constants.COLUMN_MAPPING[section_name]
        
        if level in StyleConstants.LEVEL_COLORS:
            self._fill_row_with_level_color(ws, original_row, level)
        
        for i, budget_col in enumerate(budget_columns):
            if i in zero_columns or (i + len(budget_columns)) in zero_columns:
                continue
            
            self._validate_budget_cells(ws, row_data, budget_col, i, mapping, original_row, level)
    
    def _fill_row_with_level_color(self, ws, row: int, level: int):
        """Закрашивание строки"""
        level_fill = PatternFill(
            start_color=StyleConstants.LEVEL_COLORS[level], 
            end_color=StyleConstants.LEVEL_COLORS[level], 
            fill_type="solid"
        )
        
        for col_idx in range(1, 37):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = level_fill
    
    def _validate_budget_cells(self, ws, row_data: dict, budget_col: str, col_index: int, mapping: dict, original_row: int, level: int):
        """Проверка бюджетных ячеек"""
        approved_col = mapping['утвержденный'][col_index]
        approved_col_idx = column_to_index(approved_col) + 1
        approved_cell = ws.cell(row=original_row, column=approved_col_idx)
        
        original_approved = row_data[f'утвержденный_{budget_col}']
        calculated_approved = row_data[f'расчетный_утвержденный_{budget_col}']
        
        if is_value_different(original_approved, calculated_approved) and level < 6:
            approved_cell.fill = StyleConstants.ERROR_FILL
            if self.show_error_values:
                try:
                    orig_val = float(original_approved or 0)
                except (TypeError, ValueError):
                    orig_val = 0.0
                try:
                    calc_val = float(calculated_approved or 0)
                except (TypeError, ValueError):
                    calc_val = 0.0
                approved_cell.value = f"{orig_val:.2f} ({calc_val:.2f})"
        
        executed_col = mapping['исполненный'][col_index]
        executed_col_idx = column_to_index(executed_col) + 1
        executed_cell = ws.cell(row=original_row, column=executed_col_idx)
        
        original_executed = row_data[f'исполненный_{budget_col}']
        calculated_executed = row_data[f'расчетный_исполненный_{budget_col}']
        
        if is_value_different(original_executed, calculated_executed) and level < 6:
            executed_cell.fill = StyleConstants.ERROR_FILL
            if self.show_error_values:
                try:
                    orig_val = float(original_executed or 0)
                except (TypeError, ValueError):
                    orig_val = 0.0
                try:
                    calc_val = float(calculated_executed or 0)
                except (TypeError, ValueError):
                    calc_val = 0.0
                executed_cell.value = f"{orig_val:.2f} ({calc_val:.2f})"
    
    def _hide_zero_columns(self, ws, section_name: str, zero_columns: list):
        """Скрытие нулевых столбцов"""
        mapping = self.constants.COLUMN_MAPPING[section_name]
        budget_columns = self.constants.BUDGET_COLUMNS
        
        for col_index in zero_columns:
            if col_index < len(budget_columns):
                col_letter = mapping['утвержденный'][col_index]
            else:
                col_letter = mapping['исполненный'][col_index - len(budget_columns)]
            
            col_idx = column_to_index(col_letter) + 1
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].hidden = True
    
    def _extract_original_value_from_cell(self, cell_value) -> float:
        """Извлечение оригинального значения из ячейки Excel (deprecated - используйте form_utils.extract_original_value_from_cell)"""
        return extract_original_value_from_cell(cell_value)
    
    def _is_value_different(self, original: float, calculated: float) -> bool:
        """Проверка различия значений (deprecated - используйте numeric_utils.is_value_different)"""
        return is_value_different(original, calculated)
    
    def _column_to_index(self, column_letter: str) -> int:
        """Конвертация буквы колонки в индекс (deprecated - используйте form_utils.column_to_index)"""
        return column_to_index(column_letter)
    
    def _get_cell_value(self, sheet: pd.DataFrame, row: int, col: int) -> str:
        """Получение значения ячейки (deprecated - используйте form_utils.get_cell_value)"""
        return get_cell_value(sheet, row, col)
    
    def _get_numeric_value(self, sheet: pd.DataFrame, row: int, col: int) -> float:
        """Получение числового значения (deprecated - используйте form_utils.get_numeric_value)"""
        return get_numeric_value(sheet, row, col)
    
    def _clean_dbk_code(self, code: str) -> str:
        """Очистка кода классификации (deprecated - используйте form_utils.clean_dbk_code)"""
        return clean_dbk_code(code)
    
    def _format_classification_code(self, code: str, section_type: str) -> str:
        """Форматирование кода классификации (deprecated - используйте form_utils.format_classification_code)"""
        return format_classification_code(code, section_type)