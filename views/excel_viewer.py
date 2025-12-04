from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, 
                             QTableWidgetItem, QHeaderView, QComboBox, QLabel,
                             QPushButton, QSplitter, QTabWidget, QMessageBox,
                             QMenu)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor, QBrush
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

class ExcelViewer(QWidget):
    """Виджет для просмотра Excel файлов в табличном виде"""
    
    def __init__(self):
        super().__init__()
        self.current_file_path = None
        self.workbook = None
        self.sheet_data = {}
        self.init_ui()
    
    def init_ui(self):
        """Инициализация интерфейса"""
        layout = QVBoxLayout(self)
        
        # Панель управления
        control_layout = QHBoxLayout()
        
        # Выбор листа
        control_layout.addWidget(QLabel("Лист:"))
        self.sheet_combo = QComboBox()
        self.sheet_combo.currentTextChanged.connect(self.on_sheet_changed)
        control_layout.addWidget(self.sheet_combo)
        
        # Информация
        self.info_label = QLabel("Файл не загружен")
        control_layout.addWidget(self.info_label)
        
        control_layout.addStretch()
        
        # Кнопка обновления
        self.refresh_btn = QPushButton("Обновить")
        self.refresh_btn.clicked.connect(self.refresh_data)
        control_layout.addWidget(self.refresh_btn)
        
        layout.addLayout(control_layout)
        
        # Таблица для отображения данных
        self.data_table = QTableWidget()
        self.data_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.data_table.verticalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.data_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.data_table.customContextMenuRequested.connect(self.show_context_menu)
        layout.addWidget(self.data_table)
    
    def load_excel_file(self, file_path: str):
        """Загрузка Excel файла"""
        try:
            self.current_file_path = file_path
            self.workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Заполняем список листов
            self.sheet_combo.clear()
            self.sheet_combo.addItems(self.workbook.sheetnames)
            
            # Загружаем данные всех листов
            self.sheet_data = {}
            for sheet_name in self.workbook.sheetnames:
                self.load_sheet_data(sheet_name)
            
            # Показываем первый лист
            if self.workbook.sheetnames:
                self.sheet_combo.setCurrentIndex(0)
                self.display_sheet(self.workbook.sheetnames[0])
            
            self.info_label.setText(f"Загружен: {file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить файл: {str(e)}")
    
    def load_sheet_data(self, sheet_name: str):
        """Загрузка данных листа"""
        try:
            worksheet = self.workbook[sheet_name]
            
            # Определяем размеры данных
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # Читаем данные
            data = []
            for row in range(1, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row, col)
                    value = cell.value
                    
                    # Обрабатываем разные типы данных
                    if value is None:
                        display_value = ""
                    elif isinstance(value, float):
                        display_value = f"{value:,.2f}"
                    else:
                        display_value = str(value)
                    
                    row_data.append({
                        'value': value,
                        'display_value': display_value,
                        'font_color': cell.font.color,
                        'fill_color': cell.fill.start_color if cell.fill.patternType else None,
                        'is_bold': cell.font.bold,
                        'is_italic': cell.font.italic
                    })
                data.append(row_data)
            
            self.sheet_data[sheet_name] = {
                'data': data,
                'max_row': max_row,
                'max_col': max_col
            }
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка загрузки листа {sheet_name}: {str(e)}")
    
    def display_sheet(self, sheet_name: str):
        """Отображение данных листа в таблице"""
        if sheet_name not in self.sheet_data:
            return
        
        sheet_info = self.sheet_data[sheet_name]
        data = sheet_info['data']
        max_row = sheet_info['max_row']
        max_col = sheet_info['max_col']
        
        # Настраиваем таблицу
        self.data_table.setRowCount(max_row)
        self.data_table.setColumnCount(max_col)
        
        # Устанавливаем заголовки столбцов (буквы Excel)
        column_headers = [get_column_letter(i + 1) for i in range(max_col)]
        self.data_table.setHorizontalHeaderLabels(column_headers)
        
        # Устанавливаем заголовки строк (номера Excel)
        row_headers = [str(i + 1) for i in range(max_row)]
        self.data_table.setVerticalHeaderLabels(row_headers)
        
        # Заполняем данные
        for row_idx, row_data in enumerate(data):
            for col_idx, cell_data in enumerate(row_data):
                item = QTableWidgetItem(cell_data['display_value'])
                
                # Применяем стили
                self.apply_cell_styles(item, cell_data)
                
                self.data_table.setItem(row_idx, col_idx, item)
        
        # Настраиваем размеры столбцов
        self.adjust_columns_width()
    
    def apply_cell_styles(self, item, cell_data):
        """Применение стилей ячейки"""
        # Цвет текста
        if cell_data['font_color'] and hasattr(cell_data['font_color'], 'rgb'):
            try:
                color = QColor(cell_data['font_color'].rgb)
                item.setForeground(QBrush(color))
            except:
                pass
        
        # Цвет фона
        if cell_data['fill_color'] and hasattr(cell_data['fill_color'], 'rgb'):
            try:
                color = QColor(cell_data['fill_color'].rgb)
                item.setBackground(QBrush(color))
            except:
                pass
        
        # Жирный шрифт
        if cell_data['is_bold']:
            font = item.font()
            font.setBold(True)
            item.setFont(font)
        
        # Курсив
        if cell_data['is_italic']:
            font = item.font()
            font.setItalic(True)
            item.setFont(font)
    
    def adjust_columns_width(self):
        """Автоматическая настройка ширины столбцов"""
        header = self.data_table.horizontalHeader()
        for column in range(self.data_table.columnCount()):
            header.setSectionResizeMode(column, QHeaderView.ResizeToContents)
    
    def on_sheet_changed(self, sheet_name: str):
        """Обработка смены листа"""
        if sheet_name and self.workbook:
            self.display_sheet(sheet_name)
    
    def refresh_data(self):
        """Обновление данных"""
        if self.current_file_path:
            self.load_excel_file(self.current_file_path)
    
    def show_context_menu(self, position):
        """Контекстное меню для таблицы"""
        menu = QMenu()
        
        hide_column_action = menu.addAction("Скрыть столбец")
        show_all_columns_action = menu.addAction("Показать все столбцы")
        menu.addSeparator()
        auto_resize_action = menu.addAction("Авто-размер столбцов")
        
        action = menu.exec_(self.data_table.mapToGlobal(position))
        
        if action == hide_column_action:
            self.hide_current_column()
        elif action == show_all_columns_action:
            self.show_all_columns()
        elif action == auto_resize_action:
            self.adjust_columns_width()
    
    def hide_current_column(self):
        """Скрыть текущий столбец"""
        current_column = self.data_table.currentColumn()
        if current_column >= 0:
            self.data_table.horizontalHeader().setSectionHidden(current_column, True)
    
    def show_all_columns(self):
        """Показать все столбцы"""
        for i in range(self.data_table.columnCount()):
            self.data_table.horizontalHeader().setSectionHidden(i, False)
    
    def get_current_sheet_data(self) -> list:
        """Получение данных текущего листа"""
        current_sheet = self.sheet_combo.currentText()
        if current_sheet in self.sheet_data:
            return self.sheet_data[current_sheet]['data']
        return []